import pandas as pd
import requests
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
from geopy.exc import GeocoderUnavailable
import streamlit as st
from openpyxl import load_workbook
import tempfile
import pdfplumber
import re
import os

# --- Hide Streamlit Branding and Toolbar ---
hide_streamlit_style = """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# --- Logo ---
st.image("IWG Logo (1).jpg", width=200)

# --- Load global pricing data ---
usa_data = pd.read_excel("Global Pricing.xlsx", sheet_name="USA")
canada_data = pd.read_excel("Global Pricing.xlsx", sheet_name="Canada")
market_rent_data = pd.read_excel("Global Pricing.xlsx", sheet_name="Market Rent")

def clean_columns(df):
    df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
    return df

usa_data = clean_columns(usa_data)
canada_data = clean_columns(canada_data)
market_rent_data = clean_columns(market_rent_data)
all_data = pd.concat([usa_data, canada_data], ignore_index=True)
all_data = clean_columns(all_data)

# --- Geolocator ---
geolocator = Nominatim(user_agent="pricing_app")

@st.cache_data
def get_coords_cached(address):
    """Geocode address with caching and error handling."""
    try:
        location = geolocator.geocode(address, timeout=10)
        if location:
            return (location.latitude, location.longitude)
        else:
            return None
    except GeocoderUnavailable:
        return None

def get_coords(address):
    coords = get_coords_cached(address)
    if not coords:
        st.warning("Geocoding service unavailable or address not found. You can enter coordinates manually below.")
    return coords

def format_diff(value):
    if value > 0:
        return f"{abs(value)}% higher"
    elif value < 0:
        return f"{abs(value)}% lower"
    else:
        return "Same as average"

def find_closest_comps(user_coords):
    valid_data = all_data.dropna(subset=['Latitude', 'Longitude', 'Price']).copy()
    valid_data = valid_data[(valid_data['Latitude'] != 0) & (valid_data['Longitude'] != 0)]
    valid_data['distance'] = valid_data.apply(
        lambda row: round(geodesic(user_coords, (row['Latitude'], row['Longitude'])).miles, 2),
        axis=1
    )
    sorted_data = valid_data.sort_values('distance')
    comps5 = sorted_data[['Centre #', 'Latitude', 'Longitude', 'Price', 'distance']].head(5)
    avg_price = comps5['Price'].mean()
    # Comp #1
    comp1_price = comps5.iloc[0]['Price']
    quality1 = "Higher Quality" if comp1_price > avg_price else "Lesser Quality" if comp1_price < avg_price else "Same Quality"
    diff1 = ((comp1_price - avg_price) / avg_price) * 100
    diff1_str = format_diff(round(diff1, 2))
    # Comp #2
    if len(comps5) > 1:
        comp2_price = comps5.iloc[1]['Price']
        quality2 = "Higher Quality" if comp2_price > avg_price else "Lesser Quality" if comp2_price < avg_price else "Same Quality"
        diff2 = ((comp2_price - avg_price) / avg_price) * 100
        diff2_str = format_diff(round(diff2, 2))
    else:
        quality2 = ""
        diff2_str = ""
    comp_centres = comps5['Centre #'].head(2).tolist()
    comp_distances = [f"{d} mi" for d in comps5['distance'].head(2).tolist()]
    while len(comp_centres) < 2:
        comp_centres.append("")
        comp_distances.append("")
    return comp_centres, comp_distances, quality1, quality2, diff1_str, diff2_str, avg_price

def get_avg_market_rent(comp_centres):
    rents = []
    for centre in comp_centres:
        if centre:
            matched = market_rent_data[market_rent_data['Centre #'] == centre]
            if not matched.empty and 'Market Rate' in matched.columns:
                rents.extend(matched['Market Rate'].tolist())
    if rents:
        return sum(rents) / len(rents)
    return 0.0

def find_online_coworking_osm(user_coords):
    lat, lon = user_coords
    overpass_url = "http://overpass-api.de/api/interpreter"
    radius = 10000
    step = 10000
    coworking_spaces = []
    
    while True:
        query = f"""
        [out:json];
        node["office"="coworking"](around:{radius},{lat},{lon});
        out;
        """
        response = requests.get(overpass_url, params={'data': query})
        try:
            data = response.json()
        except:
            break

        new_spaces = []
        for element in data.get('elements', []):
            name = element['tags'].get('name')
            if name:
                c_lat = element['lat']
                c_lon = element['lon']
                dist = round(geodesic(user_coords, (c_lat, c_lon)).miles, 2)
                new_spaces.append((name, dist, c_lat, c_lon))
        
        all_spaces = coworking_spaces + new_spaces
        unique_spaces = {}
        for name, dist, c_lat, c_lon in all_spaces:
            if name not in unique_spaces or dist < unique_spaces[name][0]:
                unique_spaces[name] = (dist, c_lat, c_lon)
        
        coworking_spaces = sorted(
            [(name, val[0], val[1], val[2]) for name, val in unique_spaces.items()],
            key=lambda x: x[1]
        )
        
        if len(coworking_spaces) >= 2:
            break
        radius += step

    while len(coworking_spaces) < 2:
        coworking_spaces.append(("No coworking space found", 0, None, None))
        
    return coworking_spaces[:2]

def safe_to_float(val):
    try:
        return float(val.replace(",", "")) if val and val != "." else 0.0
    except:
        return 0.0

# --- Extract Excel Data ---
def extract_from_excel(uploaded_file):
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb["10Yr Model"] if "10Yr Model" in wb.sheetnames else wb.active
        text = " ".join([str(cell.value) for row in ws.iter_rows() for cell in row if cell.value])
        currency = "USD" if "USD" in text else "CAD" if "CAD" in text else "USD"
        gross_area = market_rent = cashflow = None
        for row in ws.iter_rows(values_only=True):
            row_text = [str(x).strip().lower() if x else "" for x in row]
            if "gross area (sqft)" in " ".join(row_text):
                gross_area = next((x for x in row if isinstance(x, (int, float))), 0)
            if "market rent value" in " ".join(row_text):
                market_rent = next((x for x in row if isinstance(x, (int, float))), 0)
            if "net partner cashflow" in " ".join(row_text) and "year 1" in " ".join(row_text):
                cashflow = next((x for x in row if isinstance(x, (int, float))), 0)
        total_area = gross_area or 0
        monthly_cashflow = cashflow / 12 if cashflow else 0
        return currency, total_area, market_rent or 0, monthly_cashflow or 0
    except Exception as e:
        st.warning(f"Could not parse Excel model: {e}")
        return None

# --- Extract PDF Data ---
def extract_from_pdf(uploaded_file):
    try:
        text = ""
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        currency = "USD" if "USD" in text else "CAD" if "CAD" in text else "USD"
        total_area_match = re.search(r"Total Area Contracted.*?([\d,\.]+)", text, re.IGNORECASE)
        total_area = safe_to_float(total_area_match.group(1)) if total_area_match else 0.0
        market_rent_match = re.search(r"Market Rent Value.*?([\d,\.]+)", text, re.IGNORECASE)
        market_rent = safe_to_float(market_rent_match.group(1)) if market_rent_match else 0.0
        cashflow_match = re.search(r"Net Partner Cashflow.*?Year 1.*?([\d,\.]+)", text, re.IGNORECASE)
        cashflow = safe_to_float(cashflow_match.group(1)) if cashflow_match else 0.0
        monthly_cashflow = cashflow / 12 if cashflow else 0.0
        return currency, total_area, market_rent, monthly_cashflow
    except Exception as e:
        st.warning(f"Could not parse PDF model: {e}")
        return None

def extract_gross_area_from_pdf(uploaded_file):
    try:
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    match = re.search(r'Gross\s+Area\s+sq\s?ft\s*[:\s]\s*([\d,]+)', text, re.IGNORECASE)
                    if match:
                        return float(match.group(1).replace(',', ''))
    except Exception as e:
        st.warning(f"Error reading PDF for Gross Area: {e}")
    return None

# --- Fill Template ---
def fill_pricing_template(template_path, centre_num, centre_address, currency,
                          area_units, total_area, monthly_rent, rent_source,
                          service_charges, property_tax, comp_centres, comp_distances,
                          quality1, quality2, diff1_str, diff2_str,
                          coworking_names, coworking_distances, market_price,
                          total_cash_flow):
    if not os.path.exists(template_path):
        st.error(f"Template file not found: {template_path}")
        return None
    wb = load_workbook(template_path)
    ws = wb['Centre & Market Details']
    ws['C2'] = centre_num
    ws['C3'] = centre_address
    ws['D5'] = currency
    ws['D6'] = area_units
    ws['D7'] = total_area
    ws['D8'] = total_area * 0.5
    ws['D10'] = monthly_rent
    ws['D11'] = rent_source
    ws['D12'] = service_charges
    ws['D13'] = property_tax
    ws['D17'] = comp_centres[0]
    ws['E17'] = comp_centres[1]
    ws['D18'] = comp_distances[0]
    ws['E18'] = comp_distances[1]
    ws['D19'] = quality1
    ws['E19'] = quality2
    ws['D20'] = diff1_str
    ws['E20'] = diff2_str
    ws['D30'] = coworking_names[0] if len(coworking_names) > 0 else ""
    ws['E30'] = coworking_names[1] if len(coworking_names) > 1 else ""
    ws['D31'] = coworking_distances[0] if len(coworking_distances) > 0 else ""
    ws['E31'] = coworking_distances[1] if len(coworking_distances) > 1 else ""
    
    d10_value = ws['D10'].value or 0
    ws['D33'] = d10_value * 30
    ws['E33'] = d10_value * 30
    ws['D35'] = total_cash_flow
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp_file.name)
    return tmp_file.name

# --- Streamlit UI ---
st.title("Pricing Template 2025 Filler")
uploaded_model = st.file_uploader("Upload Financial Model (PDF/XLSX)", type=["xlsx", "xls", "pdf"])
currency = None
total_area = 0.0
monthly_rent = 0.0
total_cash_flow = 0.0

if uploaded_model:
    if uploaded_model.name.endswith(".pdf"):
        parsed = extract_from_pdf(uploaded_model)
        pdf_gross_area = extract_gross_area_from_pdf(uploaded_model)
        if pdf_gross_area:
            total_area = pdf_gross_area
    else:
        parsed = extract_from_excel(uploaded_model)
    if parsed:
        currency, total_area_parsed, monthly_rent, total_cash_flow = parsed
        total_area = total_area if total_area else total_area_parsed
        st.success("Values auto-extracted from model.")
    else:
        currency = st.selectbox("Pricing Currency", ["USD", "CAD"])
else:
    currency = st.selectbox("Pricing Currency", ["USD", "CAD"])

centre_num = st.text_input("Centre #")
centre_address = st.text_input("Centre Address")
area_units = st.selectbox("Area Units", ["SqM", "SqFt"])
rent_source = st.selectbox("Source of Market Rent", ["LL or Partner Provided", "Broker Provided or Market Report", "Benchmarked from similar centre"])
service_charges = st.number_input("Service Charges", min_value=0.0, format="%.2f")
property_tax = st.number_input("Property Tax", min_value=0.0, format="%.2f")

# --- Total Area Input ---
total_area_input = st.number_input(
    "Total Area Contracted", 
    value=float(total_area) if total_area else 0.0, 
    min_value=0.0
)

monthly_rent_override = st.number_input(
    "Override Monthly Market Rent", 
    value=float(monthly_rent) if monthly_rent else 0.0, 
    min_value=0.0
)

market_price = monthly_rent_override if monthly_rent_override > 0 else monthly_rent

if centre_address:
    user_coords = get_coords(centre_address)
    if not user_coords:
        # Allow manual entry
        lat = st.number_input("Latitude", value=0.0)
        lon = st.number_input("Longitude", value=0.0)
        user_coords = (lat, lon) if lat != 0.0 and lon != 0.0 else None

    if user_coords:
        comp_centres, comp_distances, quality1, quality2, diff1_str, diff2_str, _ = find_closest_comps(user_coords)
        market_price = get_avg_market_rent(comp_centres)
        st.markdown("### Closest Comps")
        st.write(f"**Comp #1:** {comp_centres[0]} — {comp_distances[0]} — Quality: {quality1} — {diff1_str}")
        if comp_centres[1]:
            st.write(f"**Comp #2:** {comp_centres[1]} — {comp_distances[1]} — Quality: {quality2} — {diff2_str}")
        coworking_spaces = find_online_coworking_osm(user_coords)
        coworking_names = [c[0] for c in coworking_spaces]
        coworking_distances = [f"{c[1]} mi" for c in coworking_spaces]
        st.markdown("### Nearby Coworking Spaces")
        for name, dist in zip(coworking_names, coworking_distances):
            st.write(f"**{name}** — {dist}")
    else:
        st.warning("No valid coordinates provided. Cannot display comps or coworking info.")

if st.button("Generate Pricing Template"):
    if not centre_num or not centre_address:
        st.error("Please enter Centre # and Centre Address")
    else:
        file_path = fill_pricing_template(
            "Pricing Template 2025.xlsx",
            centre_num,
            centre_address,
            currency,
            area_units,
            total_area_input,
            market_price,
            rent_source,
            service_charges,
            property_tax,
            comp_centres if centre_address else ["", ""],
            comp_distances if centre_address else ["", ""],
            quality1 if centre_address else "",
            quality2 if centre_address else "",
            diff1_str if centre_address else "",
            diff2_str if centre_address else "",
            coworking_names if centre_address else ["", ""],
            coworking_distances if centre_address else ["", ""],
            market_price,
            total_cash_flow
        )
        if file_path:
            with open(file_path, "rb") as f:
                st.download_button(
                    label="Download Filled Pricing Template",
                    data=f,
                    file_name=f"{centre_num}_Pricing_Template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
